#...
#1.导入tkinter，用于实现页面布局;
#2.选择存放有班级名单和座位信息的Excel文件并且打开;
#3.打开文件后的基本操作；
#4.打开文件后将名字取出存放到列表里，同事打开前判断是否有真实的文件；
#5.将所有座位信息取出存放到列表里并判断是否能排座位；
#6.将所有人的名字打乱顺序； --可尝试留个后门
#7.将排序后的名字依次填入座位并且写入Excel文件；
#...
from tkinter import *
#上面导入了tkinter的常规组件，比如label和button，对于特殊一点的控件要单独导入
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import openpyxl
import random
import os

#定义相应点击事件的函数
def select_file():
    global file
    #通过print首先验证点击按钮后相应成功
    print("open file")
    #先不加参数title和filetypes参数调用，然后再加上
    file = filedialog.askopenfilename(title="请选择模板文件",filetypes=[("Excel文件",'.xlsx'),])
    #global优化变量，删除文本框内的内容，先不加，内容越来越长后再加上
    entry1.delete(0,END)
    #将选择的文件路径写入到文本框内
    entry1.insert(0,file)

def random_names(names):
    #打乱顺序
    randomed_names = []
    number = len(names)

    while number !=0:
        index = random.randint(0,number - 1)
        name = names[index]
        randomed_names.append(name)
        names.pop(index)
        number = number - 1

    if(name =='李4'):
        if('李5' not in randomed_names):
            randomed_names.append('李5')
            names.remove('李5')
            number = number - 1
        else:
            randomed_names.remove('李5')
            randomed_names.append('李5')
    
    return randomed_names

file = ''
def seat():
    print("seating")
    #打开文件后将名字取出存放到列表里，同事打开前判断是否有真实的文件；
    global file #选择文件的函数中file也需要设定成global，思考为什么？
    if file =='':
        messagebox.showerror(title="错误!",message="您还没有选择模版文件哦")
    else:
        #打开文件分别获取名单和座位两个sheet
        work_book = openpyxl.load_workbook(file)
        sheet_name = work_book.worksheets[1]
        names = []
        #纪录所有名字和列表，从第二行开始
        for i in range(1,sheet_name.max_row):
            name = sheet_name.cell(i+1,2).value
            names.append(name)
        print(names)
    #找出所有的座位并纪录编号
    sheet_seat = work_book.worksheets[0]
    seats = []
    for i in range(4,sheet_seat.max_row,2):
        for j in range(1,sheet_seat.max_column+1):
            try:
                sunm = int(sheet_seat.cell(i,j).value)
                seats.append([i,j])
            except:
                print("我知道了，这",i,"行",j,"列不是座位，没事儿继续吧")
                continue
    if len(seats) < len(names):
        messagebox.showerror(title='错误',message='排不了!:'+str(len(seats))
        +'座位 VS '+str(len(seats))+'学生')
    print(seats)

    #将名字随机排序
    seated_names =  random_names(names)
    print(seated_names)

    for i in range(len(seated_names)):
        row = seats[i][0]
        col = seats[i][1]
        sheet_seat.cell(row+1,col,seated_names[i])

    np = '.'.join(os.path.basename(file).split('.')[:-1]+ ['排座位结果','xlsx'])
    out_file = os.path.dirname(file) + '/' + np
    work_book.save(out_file)

    messagebox.showinfo(title='提示',message='排座位成功！点击“确定”查看结果！')

    os.startfile(out_file)

def main():
    global entry1
    root = Tk()

    #设定一下界面的标题
    root.title("五年一班排座位软件v1.0")

    #新增一个文本控件，用于显示软件的名字
    Label1 = Label(root,text="五年一班排座位软件v1.0",font=("微软雅黑",24))
    Label2 = Label(root,text="请选择文件",font=("微软雅黑",16))
    entry1 = Entry(root, width=50)
    button1 = Button(root, command=select_file, text="点击选择文件",font=("微软雅黑",14))
    button2 = Button(root, command=seat, text="点击排座",font=("微软雅黑",18),bg="pink")


    Label1.grid(row=0,columnspan=3,pady=30)
    Label2.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    button1.grid(row=1,column=2)
    button2.grid(row=2,columnspan=3,pady=30)


    root.mainloop()

main()